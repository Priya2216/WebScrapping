import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Initialize Chrome webdriver
page_main = webdriver.Chrome()

# Read URLs from the Excel file
excel_filename = 'members_data.xlsx'
df = pd.read_excel(excel_filename)
url_list = df['URL'].tolist()

# Lists to store data
name_1 = []
location_1 = []
email_1 = []
image_1 = []
url_1 = []
place_of_birth_1=[]
DOB_1=[]
father_name_1=[]
mother_name_1=[]
Marital_Status_1=[]
Children_Details_1=[]
Profession_1=[]
Educational_1=[]
Permanent_Address_1=[]
Present_Address_1=[]
Positions_Held_1=[]
contact_number_1=[]
SpecialInterests=[]
CountriesVisited=[]
OtherInformation =[]
all_combine_details=[]

def extract_info_with_wait(container, selector):
    try:
        element = WebDriverWait(container, 60).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
        )
        return element.text
    except:
        return "N/A"
    
def extract_info_with_wait1(container, selector):
    try:
        element = WebDriverWait(container, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
        )
        return element.text
    except:
        return "N/A"

# Loop through each URL and scrape data
for url in url_list:
    page_main.get(url)
    
    time.sleep(2)

    container = WebDriverWait(page_main, 7200).until(EC.presence_of_element_located((By.XPATH, '//*[@id="mainContent"]/div/div/main/div[2]/div')))

# -------------- Profile Tab ---------------------------------------------------
    
    url_1.append(url)
    name_1.append(extract_info_with_wait(container, 'div:nth-child(1) > div > div > p.style_memberName__Srgzp'))
    location_1.append(extract_info_with_wait(container, 'div:nth-child(1) > div > div > p.style_memberParty__bbRyi'))
    email_1.append(extract_info_with_wait(container, 'div:nth-child(2) > div > div > span > div'))
    contact_number_1.append(extract_info_with_wait(container, '#mainContent > div > div > main > div.MuiPaper-root.MuiPaper-elevation.MuiPaper-rounded.MuiPaper-elevation1.MuiCard-root.mui-style-1lfm31s > div > div.MuiGrid-root.MuiGrid-container.MuiGrid-spacing-xs-2.mui-style-em6iwz > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-sm-12.MuiGrid-grid-lg-2\.5.mui-style-1qloham > div > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-9.MuiGrid-grid-sm-9.MuiGrid-grid-lg-9.mui-style-14ybvol'))
    Positions_Held_1.append(extract_info_with_wait(container, '#mainContent > div > div > main > div.MuiPaper-root.MuiPaper-elevation.MuiPaper-rounded.MuiPaper-elevation1.MuiCard-root.mui-style-1lfm31s > div > div.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-sm-12.MuiGrid-grid-lg-5\.9.mui-style-2jmuyi'))
    place_of_birth_1.append(extract_info_with_wait(container, '.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-sm-12.MuiGrid-grid-lg-6.mui-style-tletg0 > div:nth-child(2) > div:nth-child(1) > div > div > p'))

    try:
        image_element = container.find_element(By.CSS_SELECTOR, 'div:nth-child(1) > div > div > img')
        image_link = image_element.get_attribute("src")
        image_1.append(image_link)
    except:
        image_1.append("N/A")

    DOB_1.append(extract_info_with_wait(container,'.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-sm-12.MuiGrid-grid-lg-6.mui-style-tletg0 > div:nth-child(2) > div:nth-child(2) > div > div > p'))
    
    father_name_1.append(extract_info_with_wait(container,'.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-sm-12.MuiGrid-grid-lg-6.mui-style-tletg0 > div:nth-child(3) > div:nth-child(1) > div > div > p'))

    mother_name_1.append(extract_info_with_wait(container,'.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-sm-12.MuiGrid-grid-lg-6.mui-style-tletg0 > div:nth-child(3) > div:nth-child(2) > div > div > p'))

    Marital_Status_1.append(extract_info_with_wait(container,'.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-sm-12.MuiGrid-grid-lg-6.mui-style-tletg0 > div:nth-child(4) > div:nth-child(1) > div > div > p'))

    Children_Details_1.append(extract_info_with_wait(container,'.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-sm-12.MuiGrid-grid-lg-6.mui-style-tletg0 > div:nth-child(4) > div:nth-child(2) > div > div > p'))

    Profession_1.append(extract_info_with_wait(container,'.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-sm-12.MuiGrid-grid-lg-6.mui-style-tletg0 > div:nth-child(5) > div > div > div > p'))

    Educational_1.append(extract_info_with_wait(container, '.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-sm-12.MuiGrid-grid-lg-6.mui-style-tletg0 > div:nth-child(6) > div > div > div > p'))

    Permanent_Address_1.append(extract_info_with_wait(container,'#mainContent > div > div > main > div.MuiPaper-root.MuiPaper-elevation.MuiPaper-rounded.MuiPaper-elevation1.MuiCard-root.mui-style-1lfm31s > div > div.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-sm-12.MuiGrid-grid-lg-6.mui-style-tletg0 > div:nth-child(7) > div > div > div > p'))

    Present_Address_1.append(extract_info_with_wait(container, '.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-sm-12.MuiGrid-grid-lg-6.mui-style-tletg0 > div:nth-child(8) > div > div > div > p'))
    
# -------------- Other Details Tab ---------------------------------------------------

    second_tab_button = page_main.find_element(By.CSS_SELECTOR, '#mainContent > div > div > main > div.MuiPaper-root.MuiPaper-elevation.MuiPaper-rounded.MuiPaper-elevation1.MuiCard-root.mui-style-1lfm31s > div > div.MuiTabs-root.mui-style-jzpp77 > div.MuiTabs-scroller.MuiTabs-hideScrollbar.MuiTabs-scrollableX.mui-style-12qnib > div > button:nth-child(2)')
    second_tab_button.click()  
    
    time.sleep(1)

    SpecialInterests.append(extract_info_with_wait1(container,'#mainContent > div > div > main > div.MuiPaper-root.MuiPaper-elevation.MuiPaper-rounded.MuiPaper-elevation1.MuiCard-root.mui-style-1lfm31s > div > div.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div > div:nth-child(2) > div:nth-child(1) > div > div > p'))
    # print(SpecialInterests)

    CountriesVisited.append(extract_info_with_wait1(container,'#mainContent > div > div > main > div.MuiPaper-root.MuiPaper-elevation.MuiPaper-rounded.MuiPaper-elevation1.MuiCard-root.mui-style-1lfm31s > div > div.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div > div:nth-child(2) > div:nth-child(2) > div > div > p'))
    # print(CountriesVisited)
    
    OtherInformation.append(extract_info_with_wait1(container,'#mainContent > div > div > main > div.MuiPaper-root.MuiPaper-elevation.MuiPaper-rounded.MuiPaper-elevation1.MuiCard-root.mui-style-1lfm31s > div > div.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div > div:nth-child(2) > div:nth-child(3) > div > div > p'))
    # print(OtherInformation)
    
    all_combine_details.append(extract_info_with_wait1(container,'#mainContent > div > div > main > div.MuiPaper-root.MuiPaper-elevation.MuiPaper-rounded.MuiPaper-elevation1.MuiCard-root.mui-style-1lfm31s > div > div.MuiGrid-root.MuiGrid-container.mui-style-1d3bbye > div'))
    # print(all_combine_details)

# Close the webdriver
page_main.quit() 

# Create a DataFrame
data = {
    'Name': name_1,
    'Image_Link': image_1,
    'Party & State': location_1,
    'Email': email_1,
    'Contact Number': contact_number_1,
    'Place of Birth':place_of_birth_1,
    'Date of Birth':DOB_1,
    'Father’s Name':father_name_1,
    'Mother’s Name':mother_name_1,
    'Marital Status':Marital_Status_1,
    'Children Details':Children_Details_1,
    'Profession':Profession_1,
    'Educational':Educational_1,
    'Permanent Address':Permanent_Address_1,
    'Present Address':Present_Address_1,
    'Positions_Held':Positions_Held_1,
    'Website URL': url_1
}

detail_df = pd.DataFrame(data)

# Define other details data
data1 = {
    'Name': name_1,
    'Special Interests': SpecialInterests,
    'Countries Visited':CountriesVisited,
    'OtherInformation':OtherInformation,
    'all_combine_details':all_combine_details
}

data1_df = pd.DataFrame(data1)

# goverment bills

excel_filename = 'members_data.xlsx'

# Load the existing workbook
book = openpyxl.load_workbook(excel_filename)

# Delete sheet 'Sheet2' if it exists
if 'Sheet2' in book.sheetnames:
    sheet = book['Sheet2']
    book.remove(sheet)

# Write detail_df to 'Sheet2'
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    writer.book = book
    detail_df.to_excel(writer, sheet_name='Sheet2', index=False)

# Create or overwrite 'Sheet3' with data1_df
if 'Sheet3' in book.sheetnames:
    sheet = book['Sheet3']
    book.remove(sheet)

with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    writer.book = book
    data1_df.to_excel(writer, sheet_name='Sheet3', index=False)

print(f"Detail data saved to {excel_filename}, Sheet2 and Sheet3")

# Save changes and close the workbook
book.save(excel_filename)
book.close()
